# -*- coding: utf-8 -*-
"""
Created on Wed Jun  6 09:45:50 2023

@author: Joseph Montoya
"""

###############################################################################
#                   REPORTE DE SALDOS TOTALES
###############################################################################

#%% importación de librerías
import numpy as np
import pandas as pd
import os
#from datetime import datetime
from openpyxl import load_workbook

#%% FECHA DE CORTE
FECHA = 'FEBRERO-25' #servirá para el nombre del archivo

#%% IMPORTACIÓN DE ARCHIVOS

os.chdir('C:\\Users\\sanmiguel38\\Desktop\\CESAR - REPORTE SALDOS TOTALES\\2025\\febrero')

INSUMO           =    'CarteraTotalSM10032025.xlsx'
MES_PASADO       =    'SALDO_COOPACSANMIGUEL - ENERO-25_INC_CVV_DETALLADO.xlsx'
UTILIDAD_CASTIGO =    'Utilidad año castigo 2018 2019 2020 2021 2022 y 2023 - JGM para añadir a Saldos e Ingresos.xlsx'

COBRANZA         =    'Ingresos por Cobranza Febrero-25 - General.xlsx'
cobranza_hoja    =    'IngCob Feb-25'

#%%  IMPORTANDO LOS DATOS DE EXCEL  ##

#ANALIZADOR DE FECHAS
formatos = ['%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y',
            '%Y%m%d', 
            '%Y-%m-%d', 
            '%Y-%m-%d %H:%M:%S', 
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%d %H:%M:%S PM',
            '%Y-%m-%d %H:%M:%S AM',
            '%Y/%m/%d %H:%M:%S PM',
            '%Y/%m/%d %H:%M:%S AM']  # Lista de formatos a analizar

def parse_dates(date_str):
    for formato in formatos:
        try:
            return pd.to_datetime(date_str, format=formato)
        except ValueError:
            pass
    return pd.NaT

############################################################
##   1 el primero es la base de datos aún por procesar    ##
############################################################

df1 = pd.read_excel(INSUMO,
                 dtype = {'CodigoSocio'    : str, 
                          'NroDocIdentidad': str,
                          'NumeroPrestamo' : float, 
                          'NroPrestamoFC'  : object,
                          'TlfSocio'       : object, 
                          'CelularSocio'   : object,
                          'TipoCredito'    : object, 
                          'SubTipoCredito' : object}

                 ,parse_dates = ['FechaDesembolsoTXT'  #AQUI SI SE HA PROCESADO
                              # ,'FechaAsignacionAbogadoTXT'  #no procesado
                              # ,'FechaExpedienteTXT'         #no procesado
                              # ,'FechaAsignacion'            #no procesado
                              # ,'JFechaCastigo'              #no procesado
                              # ,'JFechaVentaCartera'         #no procesado
                              # ,'FechaProcesoSistemaTXT'     #no procesado
                             ],
                 date_parser = parse_dates)

#eliminación de duplicados por si acaso
df1 = df1.drop_duplicates(subset = 'NroPrestamoFC')

#eliminación de vacíos
df1.dropna(subset=['CodigoSocio', 
                   'NroDocIdentidad',
                   'NumeroPrestamo',
                   'NroPrestamoFC',
                   'TipoCredito'], inplace = True, how = 'all')
df1['NroPrestamoFC'] = df1['NroPrestamoFC'].str.strip()
############################################################
##        2 aqui va el reporte del mes pasado           ####
############################################################

df2 = pd.read_excel(MES_PASADO,
                    skiprows = 0, #aqui podrían haber cambios dependiendo de dónde están las columnas con los nombres
                    dtype    = {'NroDocIdentidad' : object,
                                'NumeroPrestamo'  : object,
                                'NroPrestamoFC'   : object})

#eliminación de duplicados por si acaso
df2 = df2.drop_duplicates(subset = 'NroPrestamoFC')

#eliminación de vacíos
df2.dropna(subset = ['NroDocIdentidad', 
                     'NumeroPrestamo',
                     'NroPrestamoFC'], inplace = True, how = 'all')
df2['NroPrestamoFC'] = df2['NroPrestamoFC'].str.strip()

############################################################
##    3 ESTE TERCER ARCHIVO ES LA COBRANZA DEL MES       ###
############################################################

df3_cobranza = pd.read_excel(COBRANZA,
                             dtype = {'codigosocio'  : object, 
                                      'doc_ident'    : object,
                                      'PagareFincore': object} ,
                             sheet_name = cobranza_hoja)
#aquí NO hay que eliminar duplicados

#eliminamos columnas vacías
df3_cobranza.dropna(subset = ['codigosocio', 
                              'doc_ident',
                              'PagareFincore'], 
                    inplace = True, 
                    how     = 'all')
df3_cobranza['PagareFincore'] = df3_cobranza['PagareFincore'].str.strip()

##########################################################################################################
#    4 el reporte 'Utilidad año castigo 2018 2019 2020 2021 Y 2022 - JGM para añadir a Saldos e Ingresos'    ##
##########################################################################################################

df4_JGM_año_castigo = pd.read_excel(UTILIDAD_CASTIGO,
                                    dtype = {'Numero Prestamo (Fox/Pond)' : object, 
                                             'Nro Prestamo Fincore'       : object})

df4_JGM_año_castigo = df4_JGM_año_castigo.drop_duplicates(subset='Nro Prestamo Fincore')

JGM_año_castigo =  df4_JGM_año_castigo[['Nro Prestamo Fincore', 'año castigo utilidad JGM']] #esta es la versión para hacer el merge
JGM_año_castigo['Nro Prestamo Fincore'] = JGM_año_castigo['Nro Prestamo Fincore'].str.strip()

#%%% copias por si acaso

dff1 = df1.copy() #dff1 será la copa de seguridad para no estar repitiendo
dff2 = df2.copy() #para no estar repitiendo el insertado de datos
dff3 = df3_cobranza.copy() #copia de seguridad

#df1 = dff1.copy()
#df2 = dff2.copy()
#df3_cobranza = dff3.copy()

#%% MERGE CON FINALIDAD DEL MES PASADO

#df1 = df1.drop(columns=['CodPrestamoTXT'])
df1.rename(columns={"Finalidad": "Finalidad TXT"}, inplace = True)

df2["CodFinalidad"] = df2["CodFinalidad"].astype(str)

df2_finalidad = df2[["NroPrestamoFC",
                     "CodFinalidad"]] #aquí revisar si tiene el nombre 'CodFinalidad' o 'Finalidad' a secas

df_resultado = df1.merge(df2_finalidad,
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["NroPrestamoFC"],
                         how      = 'left')

#%% ASIGNACIÓN CÓDIGO FINALIDAD

df_resultado['CodFinalidad'] = np.nan

def finalidad_producto(df_resultado):
    if pd.isnull(df_resultado['CodFinalidad']):
        if ('CAMPAÑA ESCOLAR' in df_resultado['Finalidad TXT']):
            return '51'
        elif ('DEPENDIENTES - RENTA 5TA' in df_resultado['Finalidad TXT']):
            return '31'
        elif ('INDEPENDIENTES - RENTA 4TA' in df_resultado['Finalidad TXT']):
            return '56'
        elif ('INDEPENDIENTES - MULTI OFICIOS' in df_resultado['Finalidad TXT']):
            return '32'
        elif (('POND' in df_resultado['OrigenPrestamo']) and \
              ('MICROEMPRESAS' in df_resultado['TipoCreditoTXT'] or ('PEQUEÑA EMPRESAS' in df_resultado['TipoCreditoTXT']))):
            return '41'
        elif (('LIBRE DISPONIBILIDAD' in df_resultado['Finalidad TXT']) and \
            ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT'])):
            return '30'
        elif (('INDEPENDIENTES - OTROS' in df_resultado['Finalidad TXT']) and \
              ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT'])):
            return '30'
        elif (('COMPRA DE PRODUCTO-BAZAR' in df_resultado['Finalidad TXT']) and \
              ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT'])):
            return '36'
        elif ('GARANTIA HIPOTECARIA' in df_resultado['Finalidad TXT']):
            return '41'
        elif (('ACTIVO FIJO' in df_resultado['Finalidad TXT']) and \
              ('PEQUEÑA EMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '15'
        elif (('CONSUMO ORDINARIO' in df_resultado['Finalidad TXT']) and \
              ('MEDIANAS EMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '19'

        # nuevoo
        elif (('MEDIANA EMPRESA - OTROS' in df_resultado['Finalidad TXT']) and \
              ('MEDIANAS EMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '19'

        elif (('CAPITAL DE TRABAJO' in df_resultado['Finalidad TXT']) and \
              ('MEDIANAS EMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '19'    
        elif (('ACTIVO FIJO' in df_resultado['Finalidad TXT']) and \
              ('MEDIANAS EMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '16'
        #no olvidar que aquí hay negaciones
        elif (('GARANTIA HIPOTECARIA' not in df_resultado['Finalidad TXT']) and \
              ('MICROEMPRESAS' in df_resultado['TipoCreditoTXT'])):
            return '25'
        #no olvidar que aquí hay negaciones
        elif ~('LIBRE DISPONIBILIDAD' in df_resultado['Finalidad TXT'] and \
               'INDEPENDIENTES - OTROS' in df_resultado['Finalidad TXT']) and \
              ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT']):
            return '34'
        elif ('COMPRA DE SERVICIOS / OTROS' in df_resultado['Finalidad TXT'] or \
              'COMPRA DE PRODUCTO-BAZAR' in df_resultado['Finalidad TXT']) and \
            ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT']):
            return '39'
        elif ('COMPRA DE PRODUCTO-BAZAR' in df_resultado['Finalidad TXT']) and \
            ('CONSUMO NO REVOLVENTE' in df_resultado['TipoCreditoTXT']):
            return '45'
        elif 'PEQUEÑA EMPRESAS' in df_resultado['TipoCreditoTXT']:
            return '15'
        elif (('POND' in df_resultado['OrigenPrestamo']) and \
              ('MICROEMPRESAS' in df_resultado['TipoCreditoTXT'] \
              or ('PEQUEÑA EMPRESAS' in df_resultado['TipoCreditoTXT']))):
            return '41'
        else:
            return 'investigar'
    else:
        return df_resultado['CodFinalidad']
    
df_resultado['CodFinalidad'] = df_resultado.apply(finalidad_producto, 
                                                  axis = 1)

#por si acasito, volvemos a corregir los 41 :v (cuanta inseguridad :'v)
def pond_41(df_resultado):
    if (('POND' in df_resultado['OrigenPrestamo']) and \
        ('MICROEMPRESAS' in df_resultado['TipoCreditoTXT'] or 'PEQUEÑA EMPRESAS' in df_resultado['TipoCreditoTXT'])):
        return '41'
    else:
        return df_resultado['CodFinalidad']
    
df_resultado['CodFinalidad'] = df_resultado.apply(pond_41, axis=1)

print('debe salir cero:')
print(df_resultado[df_resultado['CodFinalidad'] == 'investigar'].shape[0])

raro = df_resultado[df_resultado['CodFinalidad'] == 'investigar']

#%% SOLARIZANDO LOS CRÉDITOS QUE ESTÁN EN DÓLARES

df_resultado2 = df_resultado.copy()

# Crear máscara booleana que indica cuáles filas cumplen la condición

df_resultado2['MonedaTXT'] = df_resultado2['MonedaTXT'].str.strip()
mask = df_resultado2['MonedaTXT'].eq('US$')

# Seleccionar solo las filas que cumplen la condición y asignarles el resultado de la división
df_resultado2.loc[mask, 'SoloCapitalAmortizado']         = df_resultado2.loc[mask, 'SoloCapitalAmortizado']         / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'SaldoCapital']                  = df_resultado2.loc[mask, 'SaldoCapital']                  / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'InteresVencidoPactado']         = df_resultado2.loc[mask, 'InteresVencidoPactado']         / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'InteresPactadoPagado']          = df_resultado2.loc[mask, 'InteresPactadoPagado']          / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'SoloSaldoInteresVencido']       = df_resultado2.loc[mask, 'SoloSaldoInteresVencido']       / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'Saldo Deudor (Sobre la Cuota)'] = df_resultado2.loc[mask, 'Saldo Deudor (Sobre la Cuota)'] / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'InteresCompensatorioDeuda']     = df_resultado2.loc[mask, 'InteresCompensatorioDeuda']     / df_resultado2.loc[mask, 'TipoCambioTXT']
df_resultado2.loc[mask, 'InteresMoratorioDeuda']         = df_resultado2.loc[mask, 'InteresMoratorioDeuda']         / df_resultado2.loc[mask, 'TipoCambioTXT']

#con esto ya están divididas esas columnas entre el tipo de cambio si es que están en dólares
#%% SUMA DEL SALDO Y CRÉDITOS CAPITALIZADOS, o algo así

df_resultado3 = df_resultado2.copy()

def capitalizado(df_resultado3):
    if round(df_resultado3['MontoSolicitadoTXT'] - df_resultado3['SoloCapitalAmortizado'],2) < round(df_resultado3['SaldoCapital'],2):
        return 'CAPITALIZADO'
    elif round(df_resultado3['MontoSolicitadoTXT'] - df_resultado3['SoloCapitalAmortizado'],2) > round(df_resultado3['SaldoCapital'],2):
        return 'revisar'
    else:
        return ''

df_resultado4 = df_resultado3.copy()
df_resultado4['CRED CON CAPITALIZ'] = df_resultado3.apply(capitalizado, axis=1)
    
df4 = df_resultado4.copy()
mask2 = df4['CRED CON CAPITALIZ'].eq('revisar')

#REDONDEO A DOS DECIMALES
df4['MontoSolicitadoTXT']    = round(df4['MontoSolicitadoTXT'],2)
df4['SoloCapitalAmortizado'] = round(df4['SoloCapitalAmortizado'],2)
df4['SaldoCapital']          = round(df4['SaldoCapital'],2)

df4.loc[mask2, 'SaldoCapital'] = round(df4.loc[mask2, 'MontoSolicitadoTXT'] - df4.loc[mask2, 'SoloCapitalAmortizado'],2)

df4['CRED CON CAPITALIZ'] = df4['CRED CON CAPITALIZ'].replace('revisar', '')

df4['Nuevo Saldo'] = df4['Saldo Deudor (Sobre la Cuota)'] + df4['InteresCompensatorioDeuda'] + df4['InteresMoratorioDeuda']

#%% DATOS DEL MES PASADO
'siguiente fase'

#suma group by de la tabla donde está la cobranza del mes actual
df_sum_capital   = df3_cobranza.groupby('PagareFincore')['Capital'].sum().reset_index()
df_sum_INT_OTROS = df3_cobranza.groupby('PagareFincore')['Ingresos Financ'].sum().reset_index()
df_sum_INT_OTROS = df_sum_INT_OTROS.rename({'Ingresos Financ': 'Int y Otros mes actual'}, axis=1)

#esto es del mes pasado
#separo las columnas de nro de 'fincore' y 'IMPTE CASTIGADO (Asignado x PGB)' del mes pasado, para hacer un merge
df_cosas_mes_pasado = df2[['NroPrestamoFC',
                           'IMPTE CASTIGADO (Asignado x PGB)',
                           'FECHA DE CASTIGO',
                           'Capital Amortizado',
                           'Int y Otros']]

"Merge de la tabla en bruto con la de cosas del reporte del mes pasado"
df_cosas_mes_pasado = df_cosas_mes_pasado.drop_duplicates(subset = 'NroPrestamoFC')
df5 = df4.merge(df_cosas_mes_pasado,
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["NroPrestamoFC"],
                         how      = 'left')

"Merge de la tabla en bruto con el capital de la cobranza del mes actual"
df_sum_capital = df_sum_capital.drop_duplicates(subset = "PagareFincore")
df5 = df5.merge(df_sum_capital,
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["PagareFincore"],
                         how      = 'left')

"Merge de la tabla en bruto con el int de la cobranza del mes actual"
df_sum_INT_OTROS = df_sum_INT_OTROS.drop_duplicates(subset = "PagareFincore")
df5 = df5.merge(df_sum_INT_OTROS,
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["PagareFincore"],
                         how      = 'left')

'eliminando los NaN'
df5['Capital'].fillna(0, inplace = True)
df5['Int y Otros'].fillna(0, inplace = True)
df5['IMPTE CASTIGADO (Asignado x PGB)'] = df5['ImporteCastigoGG']
df5['IMPTE CASTIGADO (Asignado x PGB)'].fillna(0, inplace=True)
df5['FECHA DE CASTIGO'] = df5['FechaCastigoGG'] #probar si funciona
df5['FECHA DE CASTIGO'].fillna('--', inplace=True)
df5['Capital Amortizado'].fillna(0, inplace=True)
df5['Int y Otros mes actual'].fillna(0, inplace=True)

df6 = df5.copy()

mascara = df5['IMPTE CASTIGADO (Asignado x PGB)'] != 0

df6.loc[mascara, 'Capital Amortizado'] = df6.loc[mascara, 'Capital Amortizado'] + df6.loc[mascara, 'Capital']
df6.loc[mascara, 'Int y Otros'] = df6.loc[mascara, 'Int y Otros'] + df6.loc[mascara, 'Int y Otros mes actual']
# hasta aqui ya está sumado el capital amortizado y el 'interés' de este mes con el mes pasado,
# y filtrado según saldo castigado

#%% SALDO DEUDOR PGB

#sumar y restar
df6['Total Amortizado'] = df6['Capital Amortizado'] + df6['Int y Otros']
df6['SALDO CASTIGADO']  = df6['IMPTE CASTIGADO (Asignado x PGB)'] - df6['Total Amortizado']
df6.loc[mascara, 'SALDO DEUDOR REALISTA (SOLO PARA PGB)'] = df6.loc[mascara, 'Capital Amortizado'] + df6.loc[mascara, 'Capital']

# df6.loc[mascara, 'columna'] = df6['columna5'].where(mascara, df6['columna6'])                  
df6['booleando'] = mascara

def asignar_saldo_deudor(df6):
  if df6['IMPTE CASTIGADO (Asignado x PGB)']:  #true o diferente de cero
    return df6['SALDO CASTIGADO']
  else:
    return df6['Nuevo Saldo']

df6['SALDO DEUDOR REALISTA (SOLO PARA PGB)'] = df6.apply(asignar_saldo_deudor, axis=1)

# ya está llenado hasta esta columna SALDO DEUDOR REALISTA (SOLO PARA PGB)

#%% ASIGNACIÓN DE DATOS DEL MES PASADO

df2_observ_v_garantia = df2[['NroPrestamoFC','OBSERVACION','VALOR GARANTIA']].copy()
df2_observ_v_garantia['OBSERVACION'].fillna('--', inplace=True) #REEMPLAZANDO LOS NaN por --
df2_observ_v_garantia['VALOR GARANTIA'].fillna(0, inplace=True) #reemplazando los NaN por 0

df6 = df6.rename(columns = {'VALOR GARANTIA': 'VALOR GARANTIA ANTIGUA'})
df6 = df6.rename(columns = {'OBSERVACION'   : 'OBSERVACION ANTIGUA'})

df2_observ_v_garantia = df2_observ_v_garantia.drop_duplicates(subset = "NroPrestamoFC")
df6 = df6.merge(df2_observ_v_garantia,
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["NroPrestamoFC"],
                         how      =  'left')

#df6[['VALOR GARANTIA','OBSERVACION']]

# hasta aquí ya está el valor garantía
#%% SALDO REAL VS DEUDOR

dff6 = df6.copy()
def saldo_real_vs_deudor(dff6):
    if dff6['VALOR GARANTIA'] > 0:
        if dff6['VALOR GARANTIA'] < dff6['Nuevo Saldo']:
            return dff6['VALOR GARANTIA']
        else:
            return dff6['Nuevo Saldo']
    else:
        return dff6['Nuevo Saldo']
    
dff6['SALDO REAL (S.DEUDOR Vs. GARANTIA)'] = dff6.apply(saldo_real_vs_deudor, axis=1)

#solo para chequear
#%% ALERTA DEUDA > GARANTÍA

df7 = dff6.copy()
def alerta(df7):
    if df7['SALDO REAL (S.DEUDOR Vs. GARANTIA)'] < df7['Nuevo Saldo']:
        return "DEUDA SOBREPASA GARANTIA"
    else:
        return '--'
    
df7['ALERTA (Si Deuda sobrepasa V.Garantia)'] = df7.apply(alerta, axis=1)

# printeo de resultados solo para ver (no hay nada que corregir aunque salga alerta)
kho = df7[df7['ALERTA (Si Deuda sobrepasa V.Garantia)'] == "DEUDA SOBREPASA GARANTIA"][['SALDO REAL (S.DEUDOR Vs. GARANTIA)',
                                                                                        'VALOR GARANTIA', 
                                                                                        'Nuevo Saldo']]
# print(kho)

# print('en total hay ' + str(df7[df7['ALERTA (Si Deuda sobrepasa V.Garantia)'] == "DEUDA SOBREPASA GARANTIA"].shape[0]) + ' casos')
# print('es solo una alerta en el reporte, no hay que corregir nada realmente')

#%% ORDENAMIENTO DE COLUMNAS

df_final = df7.merge(JGM_año_castigo, 
                         left_on  = ["NroPrestamoFC"], 
                         right_on = ["Nro Prestamo Fincore"],
                         how      = 'left') #se duplicó un crédito

COLUMNAS = ['Socio',
            'CodigoSocio',
            'TipoDocumentoTXT',
            'NroDocIdentidad',
            'CodFinalidad',
            'TipoCreditoTXT',
            'NumeroPrestamo',
            'NroPrestamoFC',
            'FechaDesembolsoTXT',
            'MonedaTXT',
            'MontoSolicitadoTXT',
            'SoloCapitalAmortizado',
            'SaldoCapital',
            'CRED CON CAPITALIZ',
            'InteresVencidoPactado',
            'InteresPactadoPagado',
            'SoloSaldoInteresVencido',
            'Saldo Deudor (Sobre la Cuota)',
            'InteresCompensatorioDeuda',
            'InteresMoratorioDeuda',
            'Nuevo Saldo',
            'IMPTE CASTIGADO (Asignado x PGB)',
            'FECHA DE CASTIGO',
            'Capital Amortizado',
            'Int y Otros',
            'Total Amortizado',
            'SALDO CASTIGADO',
            'SALDO DEUDOR REALISTA (SOLO PARA PGB)',
            'OBSERVACION',
            'VALOR GARANTIA',
            'SALDO REAL (S.DEUDOR Vs. GARANTIA)',
            'ALERTA (Si Deuda sobrepasa V.Garantia)',
            'año castigo utilidad JGM',
            'FechaUltimoPagoCBTXT',
            'ImporteVencido',
            'NroCuotasVencidas',
            'NombrePlanilla',
            'Domicilio',
            'DistritoSocio',
            'TlfSocio',
            'CelularSocio',
            'EmailSocio',
            'OrigenTXT',
            'DiasVencimientoSBS',
            'Funcionaria',
            'SituacionTXT',
            'AbogadoTXT',
            'FechaAsignacionAbogadoTXT',
            'NroExpedienteTXT',
            'FechaExpedienteTXT',
            'Juzgado',
            'FechaAsignacion',
            'Etapa',
            'ObservacionAbogado',
            'JFechaCastigo',
            'JFechaVentaCartera',
            'UltObservacionSocio',
            'TipoCredito',
            'TipoCreditoTXT',
            'Finalidad TXT',
            'TipoCambioTXT',
            'FechaProcesoSistemaTXT',
            'FlagGarantiaPref',
            'OrigenPrestamo'
            ]

RESULTADO_FINAL = df_final[COLUMNAS]
df_finalizado = RESULTADO_FINAL.copy()

indice = ['InteresMoratorioDeuda','InteresCompensatorioDeuda', 'Saldo Deudor (Sobre la Cuota)', 'SoloSaldoInteresVencido',
          'InteresPactadoPagado', 'InteresVencidoPactado', 'SaldoCapital', 'SoloCapitalAmortizado']
for col in indice:
    df_finalizado.loc[:, col] = df_finalizado[col].round(2) #redondeando las columnas a dos decimales

#df_finalizado
#%% COLUMNA AUXILIAR PARA TABLAS PIVOTE
# creando las columnas auxiliares para los reportes 
#df_finalizado['Finalidad']

df_finalizado['CodFinalidad'] = pd.to_numeric(df_finalizado['CodFinalidad'], errors='coerce') #para convertir los tipos de datos a numérico
df_finalizado['CodFinalidad'].fillna(0, inplace=True) #para reemplazar los NaN por ceros

#df_finalizado.dropna(subset=['Finalidad'], inplace=True)  #####este código es peligroso porque elimina filas que tengan NaN, ni sé porqué lo puse

df_finalizado['CodFinalidad'] = df_finalizado['CodFinalidad'].astype(str)

#df_finalizado['CodFinalidad'] = df_finalizado['CodFinalidad'].str.replace('.0', '')
def asignacion_auxiliar(df_finalizado):
    if df_finalizado['CodFinalidad']   in ['15']:
        return 'PEQUEÑA EMPRESA'
    elif df_finalizado['CodFinalidad'] in ['16', '17', '18', '19']:
        return 'MEDIANA EMPRESA'   
    elif df_finalizado['CodFinalidad'] in ['30', '31', '32', '33', '56']:
        return 'LIBRE DISPONIBILIDAD.'    
    elif df_finalizado['CodFinalidad'] in ['34', '35', '36', '37', '38', '39', '51']:
        return 'CONSUMO ORD.'
    elif df_finalizado['CodFinalidad'] in ['41', '45']:
        return 'GARANTIA HIPOT (INCL HIPOCONSTR)'
    elif df_finalizado['CodFinalidad'] in ['20','21', '22', '23', '24', '25', '26', '27', '28', '29']:
        return 'MICRO COMERCIO'
    else:
        return 'investigar caso'

df_finalizado['auxiliar1'] = df_finalizado.apply(asignacion_auxiliar, 
                                                 axis = 1)

#%% REEMPLAZANDO NULOS EN VALOR GARANTIA
df_finalizado['VALOR GARANTIA'] = df_finalizado['VALOR GARANTIA'].fillna(0)

#%%
dff2_para_merge = dff2[['NroPrestamoFC', 'SaldoCapital']]
dff2_para_merge.rename(columns = {'SaldoCapital' : "SALDO CAPITAL DEL MES ANTERIOR"}, 
                       inplace = True)

df_finalizado = df_finalizado.merge(dff2_para_merge[['NroPrestamoFC',
                                                     'SALDO CAPITAL DEL MES ANTERIOR']],
                                    on  = 'NroPrestamoFC',
                                    how = 'left')

#%% VERIFICACIÓN DE DUPLICADOS

'''%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'''
### CÓDIGO PARA VERIFICAR SI HAY DUPLICADOS   ###%
###############################################%%%
# creamos una máscara booleana
mask = df_finalizado['NroPrestamoFC'].duplicated(keep=False)
df_duplicados = df_finalizado[mask]

print('NRO DE CRÉDITOS DUPLICADOS:')
print(df_duplicados.shape[0])
print('existen ' + str(df_duplicados.shape[0]) + ' duplicados')

df_finalizado = df_finalizado.drop_duplicates(subset = 'NroPrestamoFC')

#df_finalizado = data.copy()
#%% PIVOT TABLES
#creación de las tablas dinámicas para el reporte

# TABLA 11
datos_soles = df_finalizado[df_finalizado['MonedaTXT'] == 'S/']

tabla11 = datos_soles.pivot_table(#columns = 'auxiliar1',
                                      values=['SaldoCapital',
                                              'InteresVencidoPactado',
                                              'InteresPactadoPagado',
                                              'SoloSaldoInteresVencido',
                                              'Saldo Deudor (Sobre la Cuota)',
                                              'InteresCompensatorioDeuda',
                                              'InteresMoratorioDeuda',
                                              'Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = False, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )
column_order = ['SaldoCapital',
                'InteresVencidoPactado',
                'InteresPactadoPagado',
                'SoloSaldoInteresVencido',
                'Saldo Deudor (Sobre la Cuota)',
                'InteresCompensatorioDeuda',
                'InteresMoratorioDeuda',
                'Nuevo Saldo']
tabla11 = tabla11.reindex(columns = column_order).reset_index()
del column_order

total_columnas = ['auxiliar1',
                  'SaldoCapital',
                  'InteresVencidoPactado',
                  'InteresPactadoPagado',
                  'SoloSaldoInteresVencido',
                  'Saldo Deudor (Sobre la Cuota)',
                  'InteresCompensatorioDeuda',
                  'InteresMoratorioDeuda',
                  'Nuevo Saldo']


nueva_fila = pd.DataFrame([['Total',
                            tabla11['SaldoCapital'].sum().round(2),
                            tabla11['InteresVencidoPactado'].sum().round(2),
                            tabla11['InteresPactadoPagado'].sum().round(2),
                            tabla11['SoloSaldoInteresVencido'].sum().round(2),
                            tabla11['Saldo Deudor (Sobre la Cuota)'].sum().round(2),
                            tabla11['InteresCompensatorioDeuda'].sum().round(2),
                            tabla11['InteresMoratorioDeuda'].sum().round(2),
                            tabla11['Nuevo Saldo'].sum().round(2)]], columns = total_columnas)
tabla11 = pd.concat([tabla11,nueva_fila], ignore_index=True)
del total_columnas

#%%% TABLA 12

datos_12 = datos_soles[datos_soles['SituacionTXT'] == 'JUDICIAL']
tabla12 = datos_12.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )

tabla12 = tabla12.reset_index()

#%%% TABLA 13
datos_13 = datos_soles[(datos_soles['SituacionTXT'] == 'JUDICIAL NO ASIGNADO') | \
                       (datos_soles['SituacionTXT'] == 'JUDICIAL SIN EXPEDIENTE')]
tabla13 = datos_13.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )

tabla13 = tabla13.reset_index()

#%%% TABLA 14
datos_14 = datos_soles[~datos_soles['JFechaVentaCartera'].isnull()] #NO OLVIDAR QUE AQUÍ EL FILTRO ES NO NULOS
tabla14 = datos_14.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )

tabla14 = tabla14.reset_index()

#%% TABLAS EN DÓLARES
# tabla 21

df_finalizado['MonedaTXT'] = df_finalizado['MonedaTXT'].str.strip()
datos_dolares = df_finalizado[df_finalizado['MonedaTXT'] == 'US$']

tabla21 = datos_dolares.pivot_table(#columns = 'auxiliar1',
                                      values=['SaldoCapital',
                                              'InteresVencidoPactado',
                                              'InteresPactadoPagado',
                                              'SoloSaldoInteresVencido',
                                              'Saldo Deudor (Sobre la Cuota)',
                                              'InteresCompensatorioDeuda',
                                              'InteresMoratorioDeuda',
                                              'Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = False, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )
column_order = ['SaldoCapital',
                'InteresVencidoPactado',
                'InteresPactadoPagado',
                'SoloSaldoInteresVencido',
                'Saldo Deudor (Sobre la Cuota)',
                'InteresCompensatorioDeuda',
                'InteresMoratorioDeuda',
                'Nuevo Saldo']
tabla21 = tabla21.reindex(columns = column_order).reset_index()
del column_order
total_columnas = ['auxiliar1',
                  'SaldoCapital',
                  'InteresVencidoPactado',
                  'InteresPactadoPagado',
                  'SoloSaldoInteresVencido',
                  'Saldo Deudor (Sobre la Cuota)',
                  'InteresCompensatorioDeuda',
                  'InteresMoratorioDeuda',
                  'Nuevo Saldo']


nueva_fila = pd.DataFrame([['Total',
                            tabla21['SaldoCapital'].sum().round(2),
                            tabla21['InteresVencidoPactado'].sum().round(2),
                            tabla21['InteresPactadoPagado'].sum().round(2),
                            tabla21['SoloSaldoInteresVencido'].sum().round(2),
                            tabla21['Saldo Deudor (Sobre la Cuota)'].sum().round(2),
                            tabla21['InteresCompensatorioDeuda'].sum().round(2),
                            tabla21['InteresMoratorioDeuda'].sum().round(2),
                            tabla21['Nuevo Saldo'].sum().round(2)]], columns = total_columnas)
tabla21 = pd.concat([tabla21,nueva_fila], ignore_index=True)
del total_columnas

#%%% TABLA 22

datos_22 = datos_dolares[datos_dolares['SituacionTXT'] == 'JUDICIAL']
tabla22 = datos_22.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )

tabla22 = tabla22.reset_index()

#%%% TABLA 23

datos_23 = datos_dolares[(datos_dolares['SituacionTXT'] == 'JUDICIAL NO ASIGNADO') | \
                         (datos_dolares['SituacionTXT'] == 'JUDICIAL SIN EXPEDIENTE')]
tabla23 = datos_23.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )

tabla23 = tabla23.reset_index()

#%%% TABLA 24

datos_24 = datos_dolares[~datos_dolares['JFechaVentaCartera'].isnull()] #NO OLVIDAR QUE AQUÍ EL FILTRO ES NO NULOS
tabla24 = datos_24.pivot_table(#columns = 'auxiliar1',
                                      values       = ['Nuevo Saldo'], 
                                      index        = ['auxiliar1'],
                                      margins      = True, 
                                      margins_name = 'Total', #para sacar las sumatorias totales                                      
                                      aggfunc      = 'sum'
                                      )
tabla24 = tabla24.reset_index()

#%% CREACIÓN DEL EXCEL

nombre = "SALDO_COOPACSANMIGUEL - " + FECHA + "_INC_CVV_DETALLADO.xlsx"
try:
    ruta = nombre
    os.remove(ruta)
except FileNotFoundError:
    pass

df_finalizado.to_excel(nombre, 
                       index = False, 
                       engine = 'openpyxl')
print('Excel guardado')
            
# impte castigado no varía, eso solo se jala del mes anterior, a menos que haya nuevos castigos en el mes
   
#%% ESCRIBIENDO CON OPENPYXL
#añadimos las tablas pivote al final del dataframe
tabla = tabla11.copy()

num_filas = len(df_finalizado.index)
celda = 'A' + str(int(num_filas) + 5) #ya no sé para qué sirve esta vaina

#book = load_workbook("SALDO_COOPACSANMIGUEL - " + FECHA +"_INC_CVV_DETALLADO.xlsx")

book = load_workbook(nombre)
hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 5) 
columna_inicio = 1
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%%% TABLA 12
tabla = tabla12.copy()

num_filas = len(df_finalizado.index)
celda = 'L' + str(int(num_filas) + 5) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 5)
columna_inicio = 12
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]
            
#%%% TABLA 13
tabla = tabla13.copy()

num_filas = len(df_finalizado.index)
celda = 'P' + str(int(num_filas) + 5) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 5) 
columna_inicio = 16
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%%% TABLA 14
tabla = tabla14.copy()

num_filas = len(df_finalizado.index)
celda = 'T' + str(int(num_filas) + 5) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 5) 
columna_inicio = 20
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%%% TABLA 21       
tabla = tabla21.copy()

num_filas = len(df_finalizado.index)
celda = 'A' + str(int(num_filas) + 16) #ya no sé para qué sirve esta vaina

#book = load_workbook("SALDO_COOPACSANMIGUEL - " + FECHA +"_INC_CVV_DETALLADO.xlsx")

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 16) 
columna_inicio = 1
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%%% TABLA 22
tabla = tabla22.copy()

num_filas = len(df_finalizado.index)
celda = 'L' + str(int(num_filas) + 16) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 16) 
columna_inicio = 12
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]
            
#%%% TABLA 23
tabla = tabla23.copy()

num_filas = len(df_finalizado.index)
celda = 'P' + str(int(num_filas) + 16) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 16) 
columna_inicio = 16
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column=columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%%% TABLA 24
tabla = tabla24.copy()

num_filas = len(df_finalizado.index)
celda = 'T' + str(int(num_filas) + 16) #ya no sé para qué sirve esta vaina

#book = load_workbook(nombre)
#hoja = book['Sheet1']

filas, columnas = tabla.shape

fila_inicio = int(int(num_filas) + 16) 
columna_inicio = 20
# Insertar los nombres de las columnas y los datos del DataFrame en el archivo Excel
for fila in range(filas + 1):  # +1 para incluir la fila de los nombres de las columnas
    for columna in range(columnas):
        celda = hoja.cell(row=fila + fila_inicio, column = columna + columna_inicio)
        
        if fila == 0:
            # Insertar nombres de las columnas
            celda.value = tabla.columns[columna]
        else:
            # Insertar datos del DataFrame
            celda.value = tabla.iloc[fila - 1, columna]

#%% GUARDAR LOS CAMBIOS EN EL EXCEL

book.save(nombre)
book.close()

#%%

print('Fin del proceso')

